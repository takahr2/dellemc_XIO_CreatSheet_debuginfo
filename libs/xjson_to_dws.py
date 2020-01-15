import openpyxl as opxl
import os
import pathlib
import json
import shutil
import pprint

class create_xdws:

    def __init__(self, xjson):
        self.xjson = xjson

    def createdws(self, templatedws, outfolder):
        with open(self.xjson, encoding='utf-8') as j:
            xioinfo = json.load(j)
        
        clusterlist = []
        for cluster, info in xioinfo.items():
            if len(info) == 1:
                clusterlist.append(cluster)
        
        print(clusterlist)
        for cluster in clusterlist:
            dwsname = '3_' + str(cluster) + '_Design_Worksheet_v1.0.xlsx'
            targetfilepath = os.path.join(outfolder, dwsname)
            shutil.copy(templatedws, targetfilepath)

            dwsfile = os.path.join(outfolder, dwsname)
            wb = opxl.load_workbook(dwsfile)
            for clusterinfo in xioinfo[cluster]:

                baseinfo = clusterinfo['clusterinfo'][0]
                baseinfows = wb['Network情報']
                clusternamecell = baseinfows.cell(row=4, column=3)
                clusternamecell.value = baseinfo['cluster-Name']

                xioscell = baseinfows.cell(row=5, column=8)
                xioscell.value = baseinfo['SW-Version']

                serialcell = baseinfows.cell(row=6, column=8)
                serialcell.value = baseinfo['SerialNumber']
             
                volstartrow = 7
                cgstartrow = 5
                snapsetstartrow = 5
                initiatorstartrow = 5
                lunmapstartrow = 5

                for volinfo in clusterinfo['volumeinfo']:
                    volws = wb['LUN一覧']
                    
                    volstartcolumn = 2
                    volidcell = volws.cell(row=volstartrow, column=volstartcolumn)
                    volidcell.value = volinfo['index']
                    
                    volcell = volws.cell(row=volstartrow, column=volstartcolumn + 1)
                    volcell.value = volinfo['Volume-Name']
                    
                    volsizecell = volws.cell(row=volstartrow, column=volstartcolumn + 2)
                    volsizecell.value = volinfo['Vol-Size']
                    volstartrow += 1
                
                for cginfo in clusterinfo['cg']:
                    cgws = wb['CG一覧']
                    cgstartcolumn = 2

                    cgidcell = cgws.cell(row=cgstartrow, column=cgstartcolumn)
                    cgidcell.value = cginfo['index']

                    cgnamecell = cgws.cell(row=cgstartrow, column=cgstartcolumn + 1)
                    cgnamecell.value = cginfo['CG-Name']

                    for cgvolume in cginfo['Volume-List']:
                        cgvolcell = cgws.cell(row=cgstartrow, column=cgstartcolumn + 2)
                        cgvolcell.value = cgvolume
                        cgstartrow += 1
                
                for snapsetinfo in clusterinfo['snapsets']:
                    snapsetws = wb['SnapshotSet一覧']
                    snapsetstartcolumn = 2

                    snapsetidcell = snapsetws.cell(row=snapsetstartrow, column=snapsetstartcolumn)
                    snapsetidcell.value = snapsetinfo['Index']

                    snapsetnamecell = snapsetws.cell(row=snapsetstartrow, column=snapsetstartcolumn + 1)
                    snapsetnamecell.value = snapsetinfo['SnapShotSet-Name']

                    snapsetcgcell = snapsetws.cell(row=snapsetstartrow, column=snapsetstartcolumn + 2)
                    snapsetcgcell.value = snapsetinfo['Consistency-Group-Name']

                    for snapsetvolume in snapsetinfo['Volume-List']:
                        snapsetvolcell = snapsetws.cell(row=snapsetstartrow, column=snapsetstartcolumn + 3)
                        snapsetvolcell.value = snapsetvolume
                        snapsetstartrow += 1
                
                for initiatorinfo in clusterinfo['initiators']:
                    initiatorws = wb['Initiator情報']
                    initiatorstartcolumn = 2
                    
                    initiatorindexcell = initiatorws.cell(row=initiatorstartrow, column=initiatorstartcolumn)
                    initiatorindexcell.value = initiatorinfo['Index']

                    initiatornamecell = initiatorws.cell(row=initiatorstartrow, column=initiatorstartcolumn + 1)
                    initiatornamecell.value = initiatorinfo['Initiator-Name']

                    initiatorwwncell = initiatorws.cell(row=initiatorstartrow, column=initiatorstartcolumn + 5)
                    initiatorwwncell.value = initiatorinfo['Port-Address']

                    initiatorigcell = initiatorws.cell(row=initiatorstartrow, column=initiatorstartcolumn + 6)
                    initiatorigcell.value = initiatorinfo['IG-Name']

                    initiatorstartrow += 1

                for lunmapinfo in clusterinfo['lunmappings']:
                    lunmapws = wb['(XtremIO) LunMap']
                    lunmapstartcolumn = 2

                    lunmapvolnamecell = lunmapws.cell(row=lunmapstartrow, column=lunmapstartcolumn)
                    lunmapvolnamecell.value = lunmapinfo['Volume-Name']

                    lunmapigcell = lunmapws.cell(row=lunmapstartrow, column=lunmapstartcolumn + 1)
                    lunmapigcell.value = lunmapinfo['IG-Name']

                    lunmaphlucell = lunmapws.cell(row=lunmapstartrow, column=lunmapstartcolumn + 2)
                    lunmaphlucell.value = lunmapinfo['HLU']

                    lunmapstartrow += 1

            wb.save(dwsfile)
            wb.close()

           
def test():
    crrdir = pathlib.Path('.').absolute()
    jsonfile = pathlib.Path.joinpath(crrdir, 'output/1_xio.json')
    templatefile = pathlib.Path.joinpath(crrdir, 'template/3_XtremIO_Template_Design_Worksheet_v1.0.xlsx')
    outfolder = pathlib.Path.joinpath(crrdir, 'output/')
    create_xdws(jsonfile).createdws(templatefile, outfolder)
                
if __name__ == '__main__':
    test()
